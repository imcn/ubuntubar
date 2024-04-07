# 本地Chrome浏览器设置方法
from selenium import webdriver 
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

wb = load_workbook('D:\\users.xlsx')
ws = wb.active

wsmax_row = ws.max_row
wsmax_col = ws.max_column

rows = []
for row in ws.iter_rows(min_row=2,max_row=wsmax_row,min_col=1,max_col=wsmax_col,values_only=True):
    rows.append(list(row))

driver = webdriver.Chrome() 
driver.get('http://127.0.0.1/admin/') 
time.sleep(8)
#输入登录用户名和密码并登录
driver.find_element(By.ID,"user_login").send_keys('username')
time.sleep(0.5)
driver.find_element(By.ID,"user_pass").send_keys('password')
time.sleep(1)
driver.find_element(By.ID,"wp-submit").click()
time.sleep(5)
driver.find_element(By.CLASS_NAME,"dashicons-admin-users").click()
time.sleep(5)


for i in rows:
    #点击新增机构用户
    driver.find_element(By.CLASS_NAME,"page-title-action").click()
    time.sleep(3)
    #添加用户信息
    driver.find_element(By.ID,"user_login").send_keys(i[1])
    time.sleep(0.5)
    driver.find_element(By.ID,"email").send_keys(i[3])
    time.sleep(0.5)
    driver.find_element(By.ID,"first_name").send_keys(i[0])
    time.sleep(0.5)
    driver.find_element(By.ID,"pass1").clear()
    time.sleep(2)
    driver.find_element(By.ID,"pass1").send_keys(i[2])
    time.sleep(3)
    driver.find_element(By.CLASS_NAME,"pw-checkbox").click()
    time.sleep(3)
    driver.find_element(By.ID,"createusersub").click()
    time.sleep(5)
    #提交

#休息3秒
time.sleep(3)
#关闭浏览器
driver.close()