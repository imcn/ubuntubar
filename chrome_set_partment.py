# 本地Chrome浏览器设置方法
from selenium import  webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import load_workbook

#wb = load_workbook('F:\核酸检测\新增用户.xlsx')
#ws = wb.active

#wsmax_row = ws.max_row
#wsmax_col = ws.max_column

#rows = []
#for row in ws.iter_rows(min_row=2,max_row=wsmax_row,min_col=1,max_col=wsmax_col,values_only=True):
#    rows.append(list(row))

driver = webdriver.Chrome() 
driver.get('https://hsjc.scdsjzx.cn/SCHSJC/login') 
time.sleep(3)
#输入登录用户名和密码并登录
driver.find_element(By.ID,"user").send_keys('name')
time.sleep(0.5)
driver.find_element(By.ID,"password").send_keys('123456')
time.sleep(1)
driver.find_element(By.ID,"login-btn").click()
time.sleep(5)

#for i in rows:
#点击机构维护
driver.find_element(By.XPATH,"//*[@id='机构维护']/span[4]").click()
time.sleep(2)
#输入机构名称
driver.find_element(By.XPATH,"//*[@id='fieldsFilter1']/div[2]/div/span[2]/span[3]/div/input").send_keys('要录入内容')
time.sleep(0.5)
driver.find_element(By.XPATH,"//*[@id='list1']/div[2]/div/div[1]/div[4]/div/div/span/span/span").click()
time.sleep(1)
rc = driver.find_element(By.XPATH,"//*[@id='list1']/div[2]/div/div[1]/div[4]/div/div/span/span/span")
time.sleep(0.5)
ActionChains(driver).double_click(rc).perform()
time.sleep(5)
list_windows = driver.window_handles
driver.switch_to.window(list_windows[0])
#点击编辑
driver.find_element(By.XPATH,"//*[@id='button1']/div[2]/div/span[2]").click()

#休息3秒
time.sleep(3)
#关闭浏览器
driver.close()